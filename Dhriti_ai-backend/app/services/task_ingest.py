from __future__ import annotations

import ast
import json
import math
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple
from uuid import uuid4
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook

from app.schemas.task_ingest import (
    ColumnCoreConfig,
    CoreMapping,
    GenerateCoreConfig,
    MappingConfig,
    PayloadColumnSelection,
    PayloadSelection,
    PreviewIssue,
    PreviewRow,
)

INVALID_SHEET_NAME_CHARS = set('[]:*?/\\')
MAX_SHEET_NAME_LENGTH = 31


class TaskImportError(Exception):
    """Base exception for task import failures."""


class RecordsPathError(TaskImportError):
    """Raised when the provided records_path cannot be resolved."""


class TransformError(TaskImportError):
    """Raised when a transform cannot be applied."""


PREVIEW_LIMIT = 200


def sanitize_sheet_name(name: Optional[str], fallback: str = "Sheet1") -> str:
    """Normalize sheet names to satisfy Excel constraints."""
    candidate = (name or "").strip()
    base_fallback = (fallback or "Sheet1").strip() or "Sheet1"
    if not candidate:
        candidate = base_fallback
    candidate = "".join("_" if ch in INVALID_SHEET_NAME_CHARS else ch for ch in candidate)
    candidate = candidate[:MAX_SHEET_NAME_LENGTH].strip()
    if not candidate:
        candidate = "".join(
            "_" if ch in INVALID_SHEET_NAME_CHARS else ch for ch in base_fallback
        )[:MAX_SHEET_NAME_LENGTH].strip() or "Sheet1"
    return candidate


def determine_sheet_name(preferred: Optional[str], original_filename: Optional[str]) -> str:
    """Pick a safe sheet name, preferring user input else JSON filename."""
    fallback = None
    if original_filename:
        fallback = Path(original_filename).stem
    return sanitize_sheet_name(preferred, fallback or "Sheet1")


def convert_json_bytes_to_excel(
    json_bytes: bytes,
    records_path: str,
    xlsx_path: Path,
    sheet_name: str = "Raw",
) -> Tuple[List[str], int, List[List[Any]]]:
    records = load_json_records_from_bytes(json_bytes, records_path)
    flattened = [flatten_record(record) for record in records]
    columns = order_columns(flattened)
    preview_rows = build_preview_matrix(columns, flattened, limit=PREVIEW_LIMIT)
    safe_sheet_name = sanitize_sheet_name(sheet_name)
    write_excel(flattened, columns, xlsx_path, safe_sheet_name)
    return columns, len(flattened), preview_rows


def load_json_records_from_bytes(json_bytes: bytes, records_path: str) -> List[Dict[str, Any]]:
    try:
        data = json.loads(json_bytes)
    except json.JSONDecodeError as exc:
        raise TaskImportError("Invalid JSON payload.") from exc

    if not records_path or records_path.strip() in {"$", "$."}:
        return _auto_detect_records(data)

    current: Any = data
    parts = [part for part in records_path.strip().split(".") if part and part != "$"]
    try:
        for part in parts:
            if isinstance(current, list):
                index = int(part)
                current = current[index]
            elif isinstance(current, dict):
                current = current.get(part)
            else:
                raise RecordsPathError(f"Cannot descend into '{part}' from non-container type.")
    except (KeyError, ValueError, IndexError):
        raise RecordsPathError(f"records_path '{records_path}' did not resolve to an array.") from None

    if isinstance(current, list):
        return _ensure_records_list(current)
    if current is None:
        raise RecordsPathError(f"records_path '{records_path}' not found.")
    raise RecordsPathError(f"records_path '{records_path}' resolved to {type(current).__name__}, expected list.")


def _ensure_records_list(value: Iterable[Any]) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    for item in value:
        if isinstance(item, dict):
            result.append(item)
        else:
            result.append({"value": item})
    return result


def _auto_detect_records(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, list):
        return _ensure_records_list(data)
    if isinstance(data, dict):
        candidates: List[Tuple[str, List[Dict[str, Any]]]] = []
        for key, value in data.items():
            if isinstance(value, list) and value:
                dict_like = [item for item in value if isinstance(item, dict)]
                if dict_like:
                    candidates.append((key, dict_like))
        if candidates:
            _, best = max(candidates, key=lambda entry: len(entry[1]))
            return best
        return [data]
    return [{"value": data}]


def flatten_record(obj: Any, prefix: str = "") -> Dict[str, Any]:
    flattened: Dict[str, Any] = {}
    _flatten_recursive(obj, prefix, flattened)
    return flattened


def _flatten_recursive(obj: Any, prefix: str, out: Dict[str, Any]) -> None:
    if isinstance(obj, dict):
        for key, value in obj.items():
            next_key = f"{prefix}.{key}" if prefix else str(key)
            _flatten_recursive(value, next_key, out)
        return
    if isinstance(obj, list):
        if not obj:
            out[prefix or "[]"] = ""
            return
        if all(not isinstance(item, (dict, list)) for item in obj):
            out[prefix or "[]"] = "|".join(str(item) for item in obj)
        else:
            out[prefix or "[]"] = json.dumps(obj, ensure_ascii=False)
        return
    out[prefix or "value"] = obj


def order_columns(flat_rows: Sequence[Dict[str, Any]]) -> List[str]:
    if not flat_rows:
        return []
    seen: Dict[str, None] = {}
    for row in flat_rows:
        for key in row.keys():
            seen.setdefault(str(key), None)

    ordered = list(seen.keys())
    priority = [
        "id",
        "task_id",
        "taskId",
        "external_id",
        "externalId",
        "title",
        "name",
        "file",
        "file_name",
        "fileName",
        "media.url",
    ]
    prioritized = [key for key in priority if key in ordered]
    remainder = [key for key in ordered if key not in prioritized]
    return prioritized + remainder


def build_preview_matrix(
    columns: Sequence[str],
    rows: Sequence[Dict[str, Any]],
    limit: int,
) -> List[List[Any]]:
    preview: List[List[Any]] = []
    header = list(columns)
    preview.append(header)
    for row in rows[:limit]:
        preview.append([row.get(col, "") for col in columns])
    return preview


def write_excel(flat_rows: Sequence[Dict[str, Any]], columns: Sequence[str], xlsx_path: Path, sheet_name: str) -> None:
    df = pd.DataFrame(flat_rows, columns=list(columns))
    os.makedirs(xlsx_path.parent, exist_ok=True)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def read_preview_records_from_excel(
    excel_path: Path,
    sheet: Optional[str],
    limit: int,
) -> Tuple[str, List[str], List[Tuple[int, Dict[str, Any]]], int]:
    try:
        with pd.ExcelFile(excel_path) as workbook:
            sheet_names = workbook.sheet_names
            if not sheet_names:
                raise TaskImportError("Excel file has no sheets.")
            if sheet:
                if sheet not in sheet_names:
                    raise TaskImportError(
                        f"Sheet '{sheet}' not found. Available sheets: {', '.join(sheet_names)}"
                    )
                sheet_to_use = sheet
            else:
                sheet_to_use = sheet_names[0]
            df = workbook.parse(sheet_to_use, dtype=object)
    except ValueError as exc:
        raise TaskImportError(str(exc)) from exc
    except TaskImportError:
        raise
    except Exception as exc:
        raise TaskImportError(f"Unable to read Excel: {exc}") from exc

    columns = [str(col) for col in df.columns]
    total_rows = len(df.index)
    records: List[Tuple[int, Dict[str, Any]]] = []

    for offset, (_, series) in enumerate(df.head(limit).iterrows()):
        record = {columns[idx]: _sanitize_value(value) for idx, value in enumerate(series.tolist())}
        excel_row_number = offset + 2
        records.append((excel_row_number, record))

    return sheet_to_use, columns, records, total_rows


def prepare_dataset_rows(
    rows: Sequence[Dict[str, Any]],
    limit: int,
) -> Tuple[List[str], List[Tuple[int, Dict[str, Any]]], int]:
    if not rows:
        return [], [], 0

    columns: List[str] = []
    for record in rows:
        for key in record.keys():
            if key not in columns:
                columns.append(key)

    prepared: List[Tuple[int, Dict[str, Any]]] = []
    total_rows = 0

    for index, raw in enumerate(rows, start=1):
        normalized = {column: raw.get(column) for column in columns}
        if _row_is_empty(normalized.values()):
            continue
        total_rows += 1
        excel_row_number = index + 1
        if len(prepared) < limit:
            prepared.append((excel_row_number, normalized))

    return columns, prepared, total_rows


def stream_excel_rows(excel_path: Path, sheet: str) -> Iterator[Tuple[int, Dict[str, Any]]]:
    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        raise TaskImportError(f"Unable to open Excel: {exc}") from exc

    try:
        if sheet not in workbook.sheetnames:
            raise TaskImportError(f"Sheet '{sheet}' not found.")
        worksheet = workbook[sheet]
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return
        headers = _prepare_headers(header_row)
        for row_index, values in enumerate(rows_iter, start=2):
            record = {
                headers[idx]: _sanitize_value(values[idx]) if idx < len(values) else None
                for idx in range(len(headers))
            }
            if _row_is_empty(record.values()):
                continue
            yield row_index, record
    finally:
        workbook.close()


def _prepare_headers(header_row: Iterable[Any]) -> List[str]:
    headers: List[str] = []
    counts: Dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        base = str(raw).strip() if raw not in (None, "", "None") else f"column_{idx + 1}"
        if not base:
            base = f"column_{idx + 1}"
        key = base
        counter = counts.get(base, 0)
        counts[base] = counter + 1
        if counter:
            key = f"{base}_{counter + 1}"
        headers.append(key)
    return headers


def _sanitize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def _row_is_empty(values: Iterable[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


@dataclass
class MappingRuntime:
    sequence: int = 0

    def next_seq(self) -> int:
        self.sequence += 1
        return self.sequence


def process_row(
    row: Dict[str, Any],
    row_number: int,
    mapping: MappingConfig,
    runtime: MappingRuntime,
) -> Tuple[PreviewRow, List[str]]:
    issues: List[str] = []
    task_id = _resolve_task_id(row, row_number, mapping.core.task_id, runtime, issues)
    task_name = _resolve_column_value(row, row_number, mapping.core.task_name, default="Untitled", issues=issues)
    file_name = _resolve_column_value(row, row_number, mapping.core.file_name, default=f"row_{row_number - 1}.dat", issues=issues)
    payload = _build_payload(row, mapping.payload_selected, issues)
    preview_row = PreviewRow(
        row=row_number,
        task_id=str(task_id),
        task_name=str(task_name),
        file_name=str(file_name),
        payload=payload,
    )
    return preview_row, issues


def _resolve_task_id(
    row: Dict[str, Any],
    row_number: int,
    config: GenerateCoreConfig | ColumnCoreConfig,
    runtime: MappingRuntime,
    issues: List[str],
) -> str:
    if isinstance(config, GenerateCoreConfig):
        if config.strategy == "uuid_v4":
            return str(uuid4())
        if config.strategy == "seq_per_batch":
            return str(runtime.next_seq())
        if config.strategy == "expr":
            seq_value = runtime.next_seq()
            expression = (config.expression or "").strip()
            context = {
                "row_index": row_number - 2,
                "excel_row": row_number,
                "seq": seq_value,
                "row": row,
            }
            try:
                result = _safe_eval_expression(expression, context)
            except Exception as exc:
                issues.append(f"task_id expr failed → {exc}")
                return str(uuid4())
            if result in (None, "", []):
                issues.append("task_id expr empty → generated uuid")
                return str(uuid4())
            return str(result)
        issues.append(f"Unknown generation strategy: {config.strategy}")
        return str(uuid4())

    value = row.get(config.name)
    try:
        value = apply_transforms(value, config.transforms)
    except TransformError as exc:
        issues.append(f"task_id transform error → {exc}")
        value = None

    if value in (None, "", []):
        issues.append("task_id empty → generated uuid")
        return str(uuid4())
    return str(value)


def _resolve_column_value(
    row: Dict[str, Any],
    row_number: int,
    config: ColumnCoreConfig,
    *,
    default: str,
    issues: List[str],
) -> str:
    value = row.get(config.name)
    try:
        value = apply_transforms(value, config.transforms)
    except TransformError as exc:
        issues.append(f"{config.name} transform error → {exc}")
        value = None

    if value in (None, "", []):
        issues.append(f"{config.name} empty → using default")
        return default
    return str(value)


def _build_payload(
    row: Dict[str, Any],
    selections: Sequence[PayloadSelection],
    issues: List[str],
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    for selection in selections:
        if isinstance(selection, PayloadColumnSelection):
            value = row.get(selection.column)
            try:
                value = apply_transforms(value, selection.transforms)
            except TransformError as exc:
                issues.append(f"{selection.column} transform error → {exc}")
                continue
            if value in (None, "", []):
                continue
            payload[selection.key] = value
        else:
            payload[selection.key] = selection.value
    return payload


def apply_transforms(value: Any, transforms: Sequence[str]) -> Any:
    result = value
    for transform in transforms or []:
        name, args = _parse_transform(transform)
        fn = _TRANSFORMS.get(name)
        if fn is None:
            raise TransformError(f"unknown transform '{name}'")
        try:
            result = fn(result, *args)
        except Exception as exc:
            raise TransformError(str(exc)) from exc
    return result


def _parse_transform(transform: str) -> Tuple[str, List[str]]:
    text = transform.strip()
    if not text:
        raise TransformError("empty transform")
    if "(" not in text or not text.endswith(")"):
        return text.lower(), []
    name, args_str = text.split("(", 1)
    args_str = args_str[:-1]
    return name.strip().lower(), _split_args(args_str)


def _split_args(source: str) -> List[str]:
    args: List[str] = []
    buffer: List[str] = []
    quote_char = ""
    in_quote = False
    for char in source:
        if in_quote:
            if char == quote_char:
                in_quote = False
            else:
                buffer.append(char)
            continue
        if char in ("'", '"'):
            in_quote = True
            quote_char = char
            continue
        if char == ",":
            arg = "".join(buffer).strip()
            if arg:
                args.append(arg.strip("'\""))
            buffer = []
            continue
        buffer.append(char)
    tail = "".join(buffer).strip()
    if tail:
        args.append(tail.strip("'\""))
    return [arg.strip() for arg in args if arg.strip()]


def _transform_trim(value: Any, *_: str) -> Any:
    if value is None:
        return None
    return str(value).strip()


def _transform_lower(value: Any, *_: str) -> Any:
    if value is None:
        return None
    return str(value).lower()


def _transform_upper(value: Any, *_: str) -> Any:
    if value is None:
        return None
    return str(value).upper()


def _transform_to_int(value: Any, default: str = "0") -> int:
    if value in (None, ""):
        return int(default or 0)
    if isinstance(value, bool):
        return int(value)
    try:
        if isinstance(value, (int, float)):
            return int(value)
        text = str(value).strip()
        if not text:
            return int(default or 0)
        return int(float(text))
    except Exception:
        return int(default or 0)


def _transform_basename(value: Any, *_: str) -> Any:
    if value is None:
        return None
    text = str(value)
    parsed = urlparse(text)
    path = parsed.path or text
    if "/" in path:
        path = path.rsplit("/", 1)[-1]
    if "\\" in path:
        path = path.rsplit("\\", 1)[-1]
    return path or text


def _transform_split(value: Any, delimiter: str = "|") -> List[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(item) for item in value]
    text = str(value)
    if not text:
        return []
    return [part.strip() for part in text.split(delimiter)]


def _transform_join(value: Any, delimiter: str = "|") -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return delimiter.join(str(item) for item in value)
    return str(value)


_TRANSFORMS = {
    "trim": _transform_trim,
    "lower": _transform_lower,
    "upper": _transform_upper,
    "to_int": _transform_to_int,
    "to_int(0)": lambda value: _transform_to_int(value, "0"),
    "basename": _transform_basename,
    "split": _transform_split,
    "split('|')": lambda value: _transform_split(value, "|"),
    "join": _transform_join,
    "join('|')": lambda value: _transform_join(value, "|"),
}


def _safe_eval_expression(expression: str, context: Dict[str, Any]) -> Any:
    tree = ast.parse(expression, mode="eval")
    _validate_ast(tree)
    compiled = compile(tree, "<mapping_expr>", "eval")
    return eval(compiled, {"__builtins__": {}}, context)


ALLOWED_AST_NODES = (
    ast.Expression,
    ast.BinOp,
    ast.UnaryOp,
    ast.Constant,
    ast.Name,
    ast.Load,
    ast.Add,
    ast.Sub,
    ast.Mult,
    ast.Div,
    ast.Mod,
    ast.JoinedStr,
    ast.FormattedValue,
    ast.Subscript,
    ast.Slice,
    ast.Index,
    ast.Attribute,
)

ALLOWED_NAMES = {"row_index", "excel_row", "seq", "row"}


def _validate_ast(node: ast.AST) -> None:
    for child in ast.walk(node):
        if not isinstance(child, ALLOWED_AST_NODES):
            raise ValueError(f"Expression uses unsupported element '{type(child).__name__}'")
        if isinstance(child, ast.Name) and child.id not in ALLOWED_NAMES:
            raise ValueError(f"Name '{child.id}' is not allowed in expressions")
        if isinstance(child, ast.Attribute):
            raise ValueError("Attribute access is not allowed in expressions")
        if isinstance(child, ast.Subscript):
            if not isinstance(child.value, ast.Name) or child.value.id != "row":
                raise ValueError("Only row[...] subscripts are allowed")
            if isinstance(child.slice, ast.Constant):
                continue
            if isinstance(child.slice, ast.Index) and isinstance(child.slice.value, ast.Constant):
                continue
            raise ValueError("row[...] subscripts must use constant keys")


def suggest_mapping_from_columns(columns: Sequence[str]) -> MappingConfig:
    columns = list(columns)
    task_name_column = _find_column(columns, ("task_name", "name", "title"))
    if task_name_column is None and columns:
        task_name_column = columns[0]

    file_name_column = _find_column(columns, ("file_name", "filename", "file", "media.url", "url"))
    file_transforms: List[str] = []
    if file_name_column is None and columns:
        file_name_column = columns[0]
    if file_name_column and any(token in file_name_column.lower() for token in ("url", "path")):
        file_transforms.append("basename")

    core = CoreMapping(
        task_id=GenerateCoreConfig(strategy="uuid_v4"),
        task_name=ColumnCoreConfig(name=task_name_column or "task_name", transforms=["trim"]),
        file_name=ColumnCoreConfig(name=file_name_column or "file_name", transforms=file_transforms),
    )
    return MappingConfig(sheet="Raw", core=core, payload_selected=[])


def _find_column(columns: Sequence[str], hints: Sequence[str]) -> Optional[str]:
    lowered = {column.lower(): column for column in columns}
    for hint in hints:
        if hint in lowered:
            return lowered[hint]
    for column in columns:
        for hint in hints:
            if hint in column.lower():
                return column
    return None
